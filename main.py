from flask import Flask, request, send_file, jsonify, render_template
from flask_cors import CORS
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
CORS(app)

STICHTAG = datetime(2025, 1, 1)

@app.route("/")
def index():
    return render_template("index.html")

def map_geschlecht(g):
    if g == 1:
        return "w"
    elif g == 2:
        return "m"
    else:
        return "d"

def verarbeite_vereinsdatei(file):
    wb = load_workbook(file, data_only=True)
    ps = wb["Personensatz"]
    vs = wb["Vereinssatz"]

    def zeilen_als_dict(sheet):
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:] if any(row)]

    personen = zeilen_als_dict(ps)
    verein = {row[0]: row[1] for row in vs.iter_rows(min_row=2, max_col=2, values_only=True)}

    auswertung = {key: 0 for key in [
        "Aktive 7 < 18 w", "Aktive 7 < 18 m", "Aktive 7 < 18 d",
        "Aktive < 18 w", "Aktive < 18 m", "Aktive < 18 d",
        "Aktive < 27 w", "Aktive < 27 m", "Aktive < 27 d",
        "Aktive ab 18 w", "Aktive ab 18 m", "Aktive ab 18 d",
        "Aktive ab 27 w", "Aktive ab 27 m", "Aktive ab 27 d",
        "Aktive w", "Aktive m", "Aktive d", "Aktive"
    ]}

    for p in personen:
        geb = p.get("Geburtsdatum")
        if isinstance(geb, datetime):
            alter = STICHTAG.year - geb.year
        else:
            continue

        aktiv = any(p.get(k) for k in ["musikalische Früherziehung", "Stammorchester", "Jugendkapelle", "Schüler", "Senioren"])
        if not aktiv:
            continue

        g = map_geschlecht(p.get("Geschlecht"))
        auswertung[f"Aktive {g}"] += 1
        auswertung["Aktive"] += 1
        if alter < 18:
            auswertung[f"Aktive < 18 {g}"] += 1
        if alter >= 18:
            auswertung[f"Aktive ab 18 {g}"] += 1
        if alter < 27:
            auswertung[f"Aktive < 27 {g}"] += 1
        if alter >= 27:
            auswertung[f"Aktive ab 27 {g}"] += 1
        if 7 <= alter <= 17:
            auswertung[f"Aktive 7 < 18 {g}"] += 1

    def summe(prefix):
        return auswertung[f"{prefix} w"] + auswertung[f"{prefix} m"] + auswertung[f"{prefix} d"]

    daten = [{
        "Lfd. Nr.": "",
        "Verbandsnummer": verein.get("Verbandsnummer", ""),
        "Verein/Verband": verein.get("Verein", ""),
        "Titel": verein.get("Titel", ""),
        "Vorname": verein.get("Vorname", ""),
        "Name": verein.get("Name", ""),
        "Straße/Postfach": verein.get("Straße/Postfach", ""),
        "Land": "D",
        "PLZ": verein.get("PLZ", ""),
        "Ort": verein.get("Ort", ""),
        "E-Mail": verein.get("E-Mail", ""),
        "Jugendorchester": int(any(p.get("Jugendkapelle") for p in personen)),
        "Erwachsenenorchester": int(any(p.get("Stammorchester") for p in personen)),
        "Seniorenorchester": int(any(p.get("Senioren") for p in personen)),
        "Orchester gesamt": sum([
            int(any(p.get("Jugendkapelle") for p in personen)),
            int(any(p.get("Stammorchester") for p in personen)),
            int(any(p.get("Senioren") for p in personen))
        ]),
        "Aktive 7 < 18 w": auswertung["Aktive 7 < 18 w"],
        "Aktive 7 < 18 m": auswertung["Aktive 7 < 18 m"],
        "Aktive 7 < 18 d": auswertung["Aktive 7 < 18 d"],
        "Aktive 7 < 18": summe("Aktive 7 < 18"),
        "Aktive < 18 w": auswertung["Aktive < 18 w"],
        "Aktive < 18 m": auswertung["Aktive < 18 m"],
        "Aktive < 18 d": auswertung["Aktive < 18 d"],
        "Aktive < 18": summe("Aktive < 18"),
        "Aktive < 27 w": auswertung["Aktive < 27 w"],
        "Aktive < 27 m": auswertung["Aktive < 27 m"],
        "Aktive < 27 d": auswertung["Aktive < 27 d"],
        "Aktive < 27": summe("Aktive < 27"),
        "Aktive ab 18 w": auswertung["Aktive ab 18 w"],
        "Aktive ab 18 m": auswertung["Aktive ab 18 m"],
        "Aktive ab 18 d": auswertung["Aktive ab 18 d"],
        "Aktive ab 18": summe("Aktive ab 18"),
        "Aktive ab 27 w": auswertung["Aktive ab 27 w"],
        "Aktive ab 27 m": auswertung["Aktive ab 27 m"],
        "Aktive ab 27 d": auswertung["Aktive ab 27 d"],
        "Aktive ab 27": summe("Aktive ab 27"),
        "Aktive w": auswertung["Aktive w"],
        "Aktive m": auswertung["Aktive m"],
        "Aktive d": auswertung["Aktive d"],
        "Aktive": auswertung["Aktive"],
        "Fördernde": verein.get("Fördernde Mitglieder", 0)
    }]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Meldung"
    ws_out.append(list(daten[0].keys()))
    for zeile in daten:
        ws_out.append(list(zeile.values()))

    buffer = BytesIO()
    wb_out.save(buffer)
    buffer.seek(0)
    return buffer

@app.route("/upload", methods=["POST"])
def upload():
    if 'files' not in request.files:
        return jsonify({"error": "Keine Dateien empfangen"}), 400

    files = request.files.getlist("files")
    daten_gesamt = []
    for file in files:
        try:
            daten_gesamt.append(verarbeite_vereinsdatei(file))
        except Exception as e:
            return jsonify({"error": f"Fehler in Datei {file.filename}: {e}"}), 500

    return send_file(
        daten_gesamt[0],
        as_attachment=True,
        download_name=f"{datetime.today().strftime('%Y_%m_%d')}_Meldung_BVN_01.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
